import requests
from io import BytesIO
import openpyxl

# GitHub URL of the Excel workbook (replace with your actual GitHub URL)
github_url = 'https://raw.githubusercontent.com/JRKaran7/Legal-AI-Bot/main/Legal Rights.xlsx'

# Download the Excel workbook from GitHub
response = requests.get(github_url)
if response.status_code == 200:
    # Read the workbook from the response content
    workbook_data = BytesIO(response.content)

    # Load the workbook
    workbook = openpyxl.load_workbook(workbook_data, data_only=True)

    # Specify the directory where you want to save the sheets
    save_directory = 'D:\\AI Therapy Bot\\Legal\\Legal Rights'  # Replace with your desired directory

    # Iterate over all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        # Get the sheet by name
        sheet = workbook[sheet_name]

        # Create a new workbook with the current sheet
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active

        # Copy data from the original sheet to the new sheet
        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(row)

        # Save the new workbook with the current sheet
        new_workbook.save(f'{save_directory}{sheet_name}.xlsx')

    print('Sheets saved successfully.')
else:
    print(f'Failed to download the workbook. Status code: {response.status_code}')
